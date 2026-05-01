"""
简洁分析Excel文件格式
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_path = r'src\manufacturing process\Manufacturing Process - dev.xlsm'
wb = load_workbook(file_path)

print(f'=== Excel 文件格式分析: {file_path} ===\n')
print(f'Sheets数量: {len(wb.sheetnames)}')
print(f'Sheet名称: {wb.sheetnames}\n')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n--- Sheet: {sheet_name} ---')
    print(f'尺寸: 第 {ws.max_column} 列, 第 {ws.max_row} 行')
    
    # 输出第1-5行的数据
    print(f'\n前5行内容:')
    for row_idx in range(1, min(6, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(11, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            val = cell.value
            if val is None:
                row_values.append('-')
            elif isinstance(val, str) and len(val) > 20:
                row_values.append(val[:20] + '...')
            else:
                row_values.append(str(val))
        print(f'  行{row_idx}: {row_values}')
    
    # 列宽
    print(f'\n列宽配置:')
    for col_idx in range(1, min(11, ws.max_column + 1)):
        col_letter = get_column_letter(col_idx)
        width = ws.column_dimensions[col_letter].width
        if width:
            print(f'  {col_letter}: {width}')
